#!/usr/bin/env python3
"""
Auto Pseudo-Label Generator
Reads error_details_PL.xlsx and fills in blank PLs using conf/prob-based algorithm.

Excel structure per sample (8 rows):
  Row 1: dname(col1), img_id(col2), gt(col3), conf(col4), pred_chars(col5+)
  Row 2: per_char_probs(col5+)
  Row 3: aligned_pred(col5+)
  Row 4: aligned_gt(col5+)
  Row 5: 'pred'(col2), pred_string(col3)
  Row 6: 'gt'(col2),   gt_string(col3)
  Row 7: 'PL'(col2),   PL_value(col3)  <- blank if user couldn't decide
  Row 8: blank separator
"""

import argparse
import math
import openpyxl
from openpyxl.styles import Font, PatternFill


ROWS_PER_SAMPLE = 8  # 7 content rows + 1 blank


def align_sequences(s1, s2):
    """Edit distance DP with backtracking. Returns aligned_s1, aligned_s2 with '-' for gaps."""
    n, m = len(s1), len(s2)
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    for i in range(n + 1):
        dp[i][0] = i
    for j in range(m + 1):
        dp[0][j] = j
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = 1 + min(dp[i - 1][j - 1], dp[i - 1][j], dp[i][j - 1])
    aligned_s1, aligned_s2 = [], []
    i, j = n, m
    while i > 0 or j > 0:
        if i > 0 and j > 0 and (s1[i - 1] == s2[j - 1] or dp[i][j] == dp[i - 1][j - 1] + 1):
            aligned_s1.append(s1[i - 1])
            aligned_s2.append(s2[j - 1])
            i -= 1
            j -= 1
        elif i > 0 and dp[i][j] == dp[i - 1][j] + 1:
            aligned_s1.append(s1[i - 1])
            aligned_s2.append('-')
            i -= 1
        else:
            aligned_s1.append('-')
            aligned_s2.append(s2[j - 1])
            j -= 1
    return list(reversed(aligned_s1)), list(reversed(aligned_s2))


def auto_pl(pred, gt, per_char_probs, sample_conf,
            high_conf_thresh=0.85, low_conf_thresh=0.40, char_conf_thresh=0.80):
    """
    Generate pseudo label from pred/gt using confidence.

    Strategy:
      1. sample_conf >= high_conf_thresh  → trust model entirely → PL = pred
      2. sample_conf <  low_conf_thresh   → too uncertain → skip (return None)
      3. otherwise                        → per-character decision via alignment

    Per-character rules (step 3):
      - match        : take the character as-is
      - substitution : use pred_char if prob >= char_conf_thresh, else gt_char
      - insertion (pred has char, gt has '-') : keep if prob >= char_conf_thresh, else drop
      - deletion  (gt  has char, pred has '-'): keep gt_char (model missed it)

    Returns (pl_string or None, method_string)
    """
    if sample_conf >= high_conf_thresh:
        return pred, 'high-conf → pred'

    if sample_conf < low_conf_thresh:
        return None, 'low-conf → skip'

    # Per-character decision
    aligned_pred, aligned_gt = align_sequences(pred, gt)
    pred_pos = 0
    pl_chars = []

    for ap, ag in zip(aligned_pred, aligned_gt):
        if ap == ag:
            # Match: both agree
            pl_chars.append(ap)
            if ap != '-':
                pred_pos += 1
        elif ap == '-':
            # Deletion: model missed a GT char → keep GT char
            pl_chars.append(ag)
        elif ag == '-':
            # Insertion: model added a char not in GT
            prob = per_char_probs[pred_pos] if pred_pos < len(per_char_probs) else 0.0
            if prob >= char_conf_thresh:
                pl_chars.append(ap)  # confident insertion → keep
            # else: drop the inserted char
            pred_pos += 1
        else:
            # Substitution: pred_char != gt_char
            prob = per_char_probs[pred_pos] if pred_pos < len(per_char_probs) else 0.0
            if prob >= char_conf_thresh:
                pl_chars.append(ap)  # trust model
            else:
                pl_chars.append(ag)  # trust GT
            pred_pos += 1

    pl = ''.join(c for c in pl_chars if c != '-')
    return pl, f'char-level (conf={sample_conf:.4f})'


def read_samples(ws):
    """Parse all samples from the worksheet."""
    samples = []
    max_row = ws.max_row
    sample_start = 1

    while sample_start + 6 <= max_row:
        r = sample_start

        # Row 1
        dname    = ws.cell(row=r,     column=1).value
        img_id   = ws.cell(row=r,     column=2).value
        gt       = ws.cell(row=r,     column=3).value
        conf     = ws.cell(row=r,     column=4).value

        if dname is None:
            sample_start += ROWS_PER_SAMPLE
            continue

        # Pred chars from row 1, col 5+
        pred_chars = []
        col = 5
        while ws.cell(row=r, column=col).value is not None:
            pred_chars.append(str(ws.cell(row=r, column=col).value))
            col += 1
        pred = ''.join(pred_chars)

        # Per-char probs from row 2
        per_char_probs = []
        col = 5
        while ws.cell(row=r + 1, column=col).value is not None:
            per_char_probs.append(float(ws.cell(row=r + 1, column=col).value))
            col += 1

        # Row 5 col3: raw pred string (fallback if row1 chars were missing)
        raw_pred = ws.cell(row=r + 4, column=3).value
        if raw_pred is not None:
            pred = str(raw_pred)

        # Row 6 col3: raw gt string (fallback)
        raw_gt = ws.cell(row=r + 5, column=3).value
        if raw_gt is not None:
            gt = str(raw_gt)

        # Row 7 col3: user's PL (None if blank)
        user_pl = ws.cell(row=r + 6, column=3).value

        samples.append({
            'row_start': r,
            'dname':     str(dname) if dname else '',
            'img_id':    img_id,
            'gt':        str(gt) if gt else '',
            'conf':      float(conf) if conf else 0.0,
            'pred':      pred,
            'per_char_probs': per_char_probs,
            'user_pl':   str(user_pl) if user_pl is not None else None,
        })

        sample_start += ROWS_PER_SAMPLE

    return samples


def main():
    parser = argparse.ArgumentParser(description='Auto Pseudo-Label Generator')
    parser.add_argument('input_xlsx',  help='Path to error_details_PL.xlsx')
    parser.add_argument('output_xlsx', nargs='?', default=None,
                        help='Output path (default: input_auto_pl.xlsx)')
    parser.add_argument('--high-conf', type=float, default=0.85,
                        help='Sample conf threshold to fully trust pred (default: 0.85)')
    parser.add_argument('--low-conf',  type=float, default=0.40,
                        help='Sample conf threshold below which to skip (default: 0.40)')
    parser.add_argument('--char-conf', type=float, default=0.80,
                        help='Per-char prob threshold to trust pred char (default: 0.80)')
    args = parser.parse_args()

    out_path = args.output_xlsx or args.input_xlsx.replace('.xlsx', '_auto_pl.xlsx')

    wb = openpyxl.load_workbook(args.input_xlsx)
    ws = wb.active

    samples = read_samples(ws)

    font_green  = Font(color='00AA00', bold=True)
    font_orange = Font(color='FF8800', bold=True)
    font_gray   = Font(color='999999')

    stats = {'user': 0, 'auto_pred': 0, 'auto_char': 0, 'skipped': 0}

    # Process bottom-to-top so insert_rows doesn't shift subsequent samples
    for s in reversed(samples):
        r = s['row_start']

        # User PL row stays at r+6 untouched; mark green if filled
        if s['user_pl'] is not None:
            ws.cell(row=r + 6, column=3).font = font_green
            stats['user'] += 1

        pl, method = auto_pl(
            s['pred'], s['gt'], s['per_char_probs'], s['conf'],
            high_conf_thresh=args.high_conf,
            low_conf_thresh=args.low_conf,
            char_conf_thresh=args.char_conf,
        )

        # Insert a new row between user PL (r+6) and blank separator (r+7)
        ws.insert_rows(r + 7)
        auto_pl_row = r + 7

        ws.cell(row=auto_pl_row, column=2).value = 'auto_PL'
        if pl is None:
            ws.cell(row=auto_pl_row, column=2).font = font_gray
            ws.cell(row=auto_pl_row, column=3).value = None
            stats['skipped'] += 1
        else:
            ws.cell(row=auto_pl_row, column=3).value = pl
            ws.cell(row=auto_pl_row, column=3).font = font_orange
            ws.cell(row=auto_pl_row, column=4).value = method
            if 'high-conf' in method:
                stats['auto_pred'] += 1
            else:
                stats['auto_char'] += 1

    wb.save(out_path)
    print(f'Saved: {out_path}')
    print(f'\n  User-filled    : {stats["user"]}')
    print(f'  Auto (pred)    : {stats["auto_pred"]}  (conf >= {args.high_conf})')
    print(f'  Auto (char-lvl): {stats["auto_char"]}  ({args.high_conf} > conf >= {args.low_conf})')
    print(f'  Skipped        : {stats["skipped"]}  (conf < {args.low_conf})')
    print(f'  Total          : {sum(stats.values())}')


if __name__ == '__main__':
    main()
