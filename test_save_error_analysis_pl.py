#!/usr/bin/env python3
"""Error analysis script for PARSeq models trained with Pseudo-Labeling (PL) extended charset.

Identical to test_save_error_analysis.py, but replaces model.charset_adapter with
PLCharsetAdapter so that extended Unicode variant characters predicted by the PL model
are mapped back to their base characters before comparison with ground-truth labels.
"""

import argparse
import csv
import io
import json
import lmdb
import os
import re
import string
import sys
from dataclasses import dataclass
from pathlib import PurePath

from nltk import edit_distance
from PIL import Image
from tqdm import tqdm
from torchvision import transforms

import torch

from strhub.data.module import SceneTextDataModule
from strhub.data.utils import PLCharsetAdapter
from strhub.models.utils import load_from_checkpoint, parse_model_args


def sanitize_filename(s):
    """Replace characters that are invalid in filenames."""
    return s.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')


def align_sequences(s1, s2):
    """Align two sequences using edit distance DP with backtracking.
    Returns aligned_s1, aligned_s2 with '-' for gaps."""
    n, m = len(s1), len(s2)
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    for i in range(n + 1):
        dp[i][0] = i
    for j in range(m + 1):
        dp[0][j] = j
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = 1 + min(dp[i - 1][j - 1], dp[i - 1][j], dp[i][j - 1])
    # Backtrack
    aligned_s1, aligned_s2 = [], []
    i, j = n, m
    while i > 0 or j > 0:
        if i > 0 and j > 0 and (s1[i - 1] == s2[j - 1] or dp[i][j] == dp[i - 1][j - 1] + 1):
            aligned_s1.append(s1[i - 1])
            aligned_s2.append(s2[j - 1])
            i -= 1
            j -= 1
        elif i > 0 and dp[i][j] == dp[i - 1][j] + 1:
            aligned_s1.append(s1[i - 1])
            aligned_s2.append('-')
            i -= 1
        else:
            aligned_s1.append('-')
            aligned_s2.append(s2[j - 1])
            j -= 1
    return list(reversed(aligned_s1)), list(reversed(aligned_s2))


def evaluate_text(pred, gt, cased=True, punctuation=True):
    """Evaluate text with optional normalization based on flags."""
    if not cased:
        pred = pred.lower()
        gt = gt.lower()
    if not punctuation:
        pred = re.sub(r'[^\w\s]', '', pred)
        gt = re.sub(r'[^\w\s]', '', gt)
    return pred == gt, edit_distance(pred, gt) / max(len(pred), len(gt)) if max(len(pred), len(gt)) > 0 else 0


def load_original_image_from_lmdb(dataset_path, original_index):
    """Load original unprocessed image from LMDB database."""
    env = lmdb.open(dataset_path, max_readers=1, readonly=True, create=False, readahead=False, meminit=False, lock=False)
    with env.begin() as txn:
        img_key = f'image-{original_index:09d}'.encode()
        imgbuf = txn.get(img_key)
        buf = io.BytesIO(imgbuf)
        img = Image.open(buf).convert('RGB')
    env.close()
    return img


@dataclass
class Result:
    dataset: str
    num_samples: int
    accuracy: float
    ned: float
    confidence: float
    label_length: float


def print_results_table(results: list[Result], file=None):
    w = max(map(len, map(getattr, results, ['dataset'] * len(results))))
    w = max(w, len('Dataset'), len('Combined'))
    print('| {:<{w}} | # samples | Accuracy | 1 - NED | Confidence | Label Length |'.format('Dataset', w=w), file=file)
    print('|:{:-<{w}}:|----------:|---------:|--------:|-----------:|-------------:|'.format('----', w=w), file=file)
    c = Result('Combined', 0, 0, 0, 0, 0)
    for res in results:
        c.num_samples += res.num_samples
        c.accuracy += res.num_samples * res.accuracy
        c.ned += res.num_samples * res.ned
        c.confidence += res.num_samples * res.confidence
        c.label_length += res.num_samples * res.label_length
        print(
            f'| {res.dataset:<{w}} | {res.num_samples:>9} | {res.accuracy:>8.2f} | {res.ned:>7.2f} '
            f'| {res.confidence:>10.2f} | {res.label_length:>12.2f} |',
            file=file,
        )
    c.accuracy /= c.num_samples
    c.ned /= c.num_samples
    c.confidence /= c.num_samples
    c.label_length /= c.num_samples
    print('|-{:-<{w}}-|-----------|----------|---------|------------|--------------|'.format('----', w=w), file=file)
    print(
        f'| {c.dataset:<{w}} | {c.num_samples:>9} | {c.accuracy:>8.2f} | {c.ned:>7.2f} '
        f'| {c.confidence:>10.2f} | {c.label_length:>12.2f} |',
        file=file,
    )


@torch.inference_mode()
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('checkpoint', help="Model checkpoint (or 'pretrained=<model_id>')")
    parser.add_argument('--data_root', default='/data/isaackang/data/STR/parseq/english/lmdb')
    parser.add_argument('--batch_size', type=int, default=512)
    parser.add_argument('--num_workers', type=int, default=4)
    parser.add_argument('--cased', action='store_true', default=False, help='Cased comparison')
    parser.add_argument('--punctuation', action='store_true', default=False, help='Check punctuation')
    parser.add_argument('--new', action='store_true', default=False, help='Evaluate on new benchmark datasets')
    parser.add_argument('--rotation', type=int, default=0, help='Angle of rotation (counter clockwise) in degrees.')
    parser.add_argument('--device', default='cuda')
    parser.add_argument('--unsolvable_csv', type=str, default=None,
                        help='Path to CSV with label_noise/illegible entries (dataset_name, image_id, reason)')
    parser.add_argument('--save_images', action='store_true', default=False,
                        help='Save error/unsolvable/correct images to disk')
    parser.add_argument(
        '--unicode_mapping',
        default='confusion_pl_output/unicode_mapping.json',
        help='Path to unicode_mapping.json produced by confusion_and_pl.py',
    )
    args, unknown = parser.parse_known_args()
    kwargs = parse_model_args(unknown)

    # Always use full charset: digits + lowercase + uppercase + punctuation
    charset_test = string.digits + string.ascii_lowercase + string.ascii_uppercase + string.punctuation
    kwargs.update({'charset_test': charset_test})
    print(f'Using full charset with all characters')
    print(f'Evaluation mode: cased={args.cased}, punctuation={args.punctuation}')
    print(f'Additional keyword arguments: {kwargs}')

    # Load unicode_mapping.json and build ext_to_base mapping
    with open(args.unicode_mapping, 'r', encoding='utf-8') as f:
        unicode_mapping = json.load(f)
    ext_to_base = {v['unicode']: v['base_char'] for v in unicode_mapping.values()}
    print(f'Loaded unicode mapping: {len(ext_to_base)} extended chars from {args.unicode_mapping}')

    # Load unsolvable sample set if provided (maps (dataset_name, image_id) -> reason)
    unsolvable_map = {}
    if args.unsolvable_csv:
        with open(args.unsolvable_csv, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                unsolvable_map[(row['dataset_name'], int(row['image_id']))] = row['reason']
        print(f'Loaded {len(unsolvable_map)} unsolvable entries from {args.unsolvable_csv}')

    model = load_from_checkpoint(args.checkpoint, **kwargs).eval().to(args.device)
    # Replace charset_adapter with PLCharsetAdapter to map extended chars back to base chars
    model.charset_adapter = PLCharsetAdapter(charset_test, ext_to_base)

    hp = model.hparams
    datamodule = SceneTextDataModule(
        args.data_root,
        '_unused_',
        hp.img_size,
        hp.max_label_length,
        hp.charset_train,
        hp.charset_test,
        args.batch_size,
        args.num_workers,
        False,
        rotation=args.rotation,
    )

    test_set = SceneTextDataModule.TEST_BENCHMARK_SUB
    test_set = sorted(set(test_set))

    # Create output directory structure (relative to script location)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'str_error_analysis')
    os.makedirs(output_dir, exist_ok=True)
    for dataset_name in test_set:
        dataset_dir = os.path.join(output_dir, dataset_name)
        os.makedirs(dataset_dir, exist_ok=True)
        print(f'Created directory: {dataset_dir}')

    results = {}
    all_errors = []  # Collect all errors across datasets
    all_preds = []   # Collect all (pred, gt) pairs across datasets
    # Confidence lists by category (3-way: unsolvable reason)
    conf_label_noise = []
    conf_illegible = []
    conf_normal = []
    # Confidence lists by category (2-way: correct/error)
    conf_correct = []
    conf_error = []
    error_details = []  # (sample_conf, name, sample_idx, pred, gt, per_char_probs)
    max_width = max(map(len, test_set))
    for name, dataloader in datamodule.test_dataloaders(test_set).items():
        total = 0
        correct = 0
        ned = 0
        confidence = 0
        label_length = 0

        # Get dataset info for loading original images
        dataset = dataloader.dataset
        dataset_path = dataset.root
        filtered_indices = dataset.filtered_index_list

        # Prepare output files for this dataset
        output_file = os.path.join(output_dir, name, 'results.txt')
        errors_file = os.path.join(output_dir, name, 'errors.txt')
        errors_image_dir = os.path.join(output_dir, name, 'error_images')
        os.makedirs(errors_image_dir, exist_ok=True)

        with open(output_file, 'w') as f, open(errors_file, 'w') as f_err:
            sample_idx = 0
            for imgs, labels in tqdm(iter(dataloader), desc=f'{name:>{max_width}}'):
                # Get model predictions
                imgs_device = imgs.to(model.device)
                logits = model.forward(imgs_device)
                probs = logits.softmax(-1)
                preds, probs_list = model.tokenizer.decode(probs)

                # Process each sample in the batch
                for img_idx, (pred, prob, gt) in enumerate(zip(preds, probs_list, labels)):
                    image_name = f'{name}_{sample_idx:06d}'
                    image_id_only = f'{sample_idx:06d}'  # Just the numeric part
                    pred = model.charset_adapter(pred)

                    all_preds.append((pred, gt))
                    # Write to results file
                    f.write(f'{image_name}, {pred}, {gt}\n')
                    f.flush()

                    # Evaluate based on flags (model has full charset, but evaluation respects flags)
                    is_correct, ned_value = evaluate_text(pred, gt, cased=args.cased, punctuation=args.punctuation)

                    # Write to errors file if incorrect (save unprocessed text and image)
                    if not is_correct:
                        f_err.write(f'{image_name}, {pred}, {gt}\n')
                        f_err.flush()

                        # Collect error for summary (use image_id without dataset name)
                        all_errors.append((name, image_id_only, pred, gt))

                        # Collect per-character probs for error detail (exclude EOS prob)
                        per_char_probs = prob[:len(pred)].tolist()
                        sample_conf_val = prob.log().mean().exp().item()
                        error_details.append((sample_conf_val, name, sample_idx, pred, gt, per_char_probs))

                        # Save original unprocessed error image
                        if args.save_images:
                            original_index = filtered_indices[sample_idx]
                            img_original = load_original_image_from_lmdb(dataset_path, original_index)
                            img_filename = os.path.join(errors_image_dir, f'{image_name}.png')
                            img_original.save(img_filename)

                    # Categorize confidence
                    if unsolvable_map:
                        sample_conf = prob.log().mean().exp().item()
                        original_index = filtered_indices[sample_idx]
                        entry = (sample_conf, name, sample_idx, pred, gt, dataset_path, original_index)
                        # 3-way: label_noise / illegible / normal
                        reason = unsolvable_map.get((name, sample_idx))
                        if reason == 'label_noise':
                            conf_label_noise.append(entry)
                        elif reason == 'illegible':
                            conf_illegible.append(entry)
                        else:
                            conf_normal.append(entry)
                        # 2-way: correct / error
                        if is_correct:
                            conf_correct.append(entry)
                        else:
                            conf_error.append(entry)

                    # Update statistics
                    confidence += prob.prod().item()
                    if is_correct:
                        correct += 1
                    ned += ned_value
                    total += 1
                    label_length += len(pred)
                    sample_idx += 1

        accuracy = 100 * correct / total
        mean_ned = 100 * (1 - ned / total)
        mean_conf = 100 * confidence / total
        mean_label_length = label_length / total
        results[name] = Result(name, total, accuracy, mean_ned, mean_conf, mean_label_length)

    # Write error summary CSV file
    error_summary_file = os.path.join(output_dir, 'error_summary.csv')
    with open(error_summary_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['dataset_name', 'image_id', 'pred', 'gt'])
        writer.writerows(all_errors)
    print(f'Error summary saved to: {error_summary_file}')

    # Write all pred/gt pairs
    all_preds_file = os.path.join(output_dir, 'all_preds.txt')
    with open(all_preds_file, 'w') as f:
        for pred, gt in all_preds:
            f.write(f'{pred}\t{gt}\n')
    print(f'All preds saved to: {all_preds_file}')

    # Write error pred/gt pairs
    error_preds_file = os.path.join(output_dir, 'error_preds.txt')
    with open(error_preds_file, 'w') as f:
        for dname, img_id, pred, gt in all_errors:
            f.write(f'{pred}\t{gt}\n')
    print(f'Error preds saved to: {error_preds_file}')

    # Write error pred/gt pairs sorted by conf ascending
    error_preds_sorted_file = os.path.join(output_dir, 'error_preds_sorted.txt')
    if error_details:
        sorted_details = sorted(error_details, key=lambda e: e[0])
        with open(error_preds_sorted_file, 'w') as f:
            for sample_conf_val, dname, img_id, pred, gt, per_char_probs in sorted_details:
                f.write(f'{pred}\t{gt}\n')
        print(f'Error preds (sorted by conf) saved to: {error_preds_sorted_file}')

    # Write error details Excel (per-character prob)
    if error_details:
        import openpyxl
        from openpyxl.styles import Font
        font_red = Font(color='FF0000')
        font_blue = Font(color='0000FF')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'error_details'
        row_num = 1
        error_details.sort(key=lambda e: e[0])
        for sample_conf_val, dname, img_id, pred, gt, per_char_probs in error_details:
            # Row 1: dataset_name, image_id, gt, conf, pred chars one per cell
            ws.cell(row=row_num, column=1, value=dname)
            ws.cell(row=row_num, column=2, value=img_id)
            ws.cell(row=row_num, column=3, value=gt)
            ws.cell(row=row_num, column=4, value=round(sample_conf_val, 4))
            for ci, ch in enumerate(pred):
                ws.cell(row=row_num, column=5 + ci, value=ch)
            # Row 2: probs under each pred char
            for ci, p in enumerate(per_char_probs):
                ws.cell(row=row_num + 1, column=5 + ci, value=round(p, 4))
            # Row 3-4: edit distance aligned pred and gt with color
            aligned_pred, aligned_gt = align_sequences(pred, gt)
            for ci, (ap, ag) in enumerate(zip(aligned_pred, aligned_gt)):
                cell_pred = ws.cell(row=row_num + 2, column=5 + ci, value=ap)
                cell_gt = ws.cell(row=row_num + 3, column=5 + ci, value=ag)
                if ap == '-' or ag == '-':
                    cell_pred.font = font_blue
                    cell_gt.font = font_blue
                elif ap != ag:
                    cell_pred.font = font_red
                    cell_gt.font = font_red
            # Row 5: raw pred string (col 3, same column as gt)
            ws.cell(row=row_num + 4, column=2, value='pred')
            ws.cell(row=row_num + 4, column=3, value=pred)
            # Row 6: raw gt string (col 3)
            ws.cell(row=row_num + 5, column=2, value='gt')
            ws.cell(row=row_num + 5, column=3, value=gt)
            # Row 7: empty for user's PL input
            ws.cell(row=row_num + 6, column=2, value='PL')
            row_num += 8  # 7 rows + 1 blank row
        correct_error_dir = os.path.join(output_dir, 'correct_error')
        os.makedirs(correct_error_dir, exist_ok=True)
        excel_path = os.path.join(correct_error_dir, 'error_details.xlsx')
        wb.save(excel_path)
        print(f'Error details Excel saved to: {excel_path}')

    if unsolvable_map:
        print(f'\n=== Confidence Comparison (geometric mean) ===')
        print(f'\n  [2-way: Correct / Error]')
        for label, entries in [('Correct', conf_correct),
                               ('Error', conf_error)]:
            if entries:
                confs = [e[0] for e in entries]
                mean_c = sum(confs) / len(confs)
                min_c = min(confs)
                max_c = max(confs)
                print(f'  {label:<15} | count: {len(confs):>5} | mean: {mean_c:.4f} | min: {min_c:.4f} | max: {max_c:.4f}')
            else:
                print(f'  {label:<15} | count:     0 | mean:   N/A | min:   N/A | max:   N/A')

        print(f'\n  [3-way: Label Noise / Illegible / Normal]')
        for label, entries in [('Label Noise', conf_label_noise),
                               ('Illegible', conf_illegible),
                               ('Normal', conf_normal)]:
            if entries:
                confs = [e[0] for e in entries]
                mean_c = sum(confs) / len(confs)
                min_c = min(confs)
                max_c = max(confs)
                print(f'  {label:<15} | count: {len(confs):>5} | mean: {mean_c:.4f} | min: {min_c:.4f} | max: {max_c:.4f}')
            else:
                print(f'  {label:<15} | count:     0 | mean:   N/A | min:   N/A | max:   N/A')

        # Full list sorted by confidence (high to low) for each unsolvable category
        for label, folder, entries in [('Label Noise', 'label_noise', conf_label_noise),
                                       ('Illegible', 'illegible', conf_illegible)]:
            if entries:
                sorted_entries = sorted(entries, key=lambda e: e[0], reverse=True)
                print(f'\n  --- {label} (all, sorted by confidence desc) ---')
                print(f'  {"dataset_name":<15} {"image_id":>8} {"pred":<20} {"gt":<20} {"conf":>8}')
                for conf, dname, img_id, pred, gt, ds_path, orig_idx in sorted_entries:
                    print(f'  {dname:<15} {img_id:>8} {pred:<20} {gt:<20} {conf:>8.4f}')

                if args.save_images:
                    unsolvable_dir = os.path.join(output_dir, 'unsolvable')
                    img_dir = os.path.join(unsolvable_dir, folder)
                    os.makedirs(img_dir, exist_ok=True)
                    for conf, dname, img_id, pred, gt, ds_path, orig_idx in sorted_entries:
                        img = load_original_image_from_lmdb(ds_path, orig_idx)
                        img_filename = sanitize_filename(f'{conf:.4f}_{dname}_{img_id}_pred[{pred}]_gt[{gt}].png')
                        img.save(os.path.join(img_dir, img_filename))
                    print(f'  Saved {len(entries)} {label.lower()} images to {img_dir}')

        # Save images for correct/error categories
        if args.save_images:
            correct_error_dir = os.path.join(output_dir, 'correct_error')
            for label, folder, entries in [('Correct', 'correct', conf_correct),
                                           ('Error', 'error', conf_error)]:
                if entries:
                    img_dir = os.path.join(correct_error_dir, folder)
                    os.makedirs(img_dir, exist_ok=True)
                    sorted_entries = sorted(entries, key=lambda e: e[0], reverse=True)
                    for conf, dname, img_id, pred, gt, ds_path, orig_idx in sorted_entries:
                        img = load_original_image_from_lmdb(ds_path, orig_idx)
                        img_filename = sanitize_filename(f'{conf:.4f}_{dname}_{img_id}_pred[{pred}]_gt[{gt}].png')
                        img.save(os.path.join(img_dir, img_filename))
                    print(f'  Saved {len(entries)} {label.lower()} images to {img_dir}')

    result_groups = {
        'Benchmark (Subset)': SceneTextDataModule.TEST_BENCHMARK_SUB,
    }
    with open(args.checkpoint + '.log.txt', 'w') as f:
        for out in [f, sys.stdout]:
            for group, subset in result_groups.items():
                print(f'{group} set:', file=out)
                print_results_table([results[s] for s in subset], out)
                print('\n', file=out)


if __name__ == '__main__':
    main()
