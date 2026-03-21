#!/usr/bin/env python3
"""
Extract samples without user-filled PL from error_details_PL.xlsx.
- Saves a filtered Excel file (same format, only blank-PL entries)
- Copies corresponding images from error/ to error_illegible/
"""

import os
import re
import shutil
import openpyxl
from openpyxl.styles import Font


ROWS_PER_SAMPLE = 8  # rows 1-7 content + 1 blank
BASE_DIR = os.path.join('str_error_analysis', 'correct_error')
INPUT_XLSX = os.path.join(BASE_DIR, 'error_details_PL.xlsx')
ERROR_IMG_DIR = os.path.join(BASE_DIR, 'error')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'error_details_illegible.xlsx')
OUTPUT_TXT = os.path.join(BASE_DIR, 'error_illegible_preds.txt')
OUTPUT_TSV = os.path.join(BASE_DIR, 'error_illegible_pred_gt.txt')
OUTPUT_IMG_DIR = os.path.join(BASE_DIR, 'error_illegible')


def find_image(error_img_dir, dname, img_id, pred, gt):
    """Find matching image file by dataset_name and image_id."""
    img_id_str = str(int(img_id)) if isinstance(img_id, float) else str(img_id)
    for fname in os.listdir(error_img_dir):
        # Pattern: {conf}_{dname}_{img_id}_pred[...]_gt[...].png
        if f'_{dname}_{img_id_str}_' in fname:
            return fname
    return None


def main():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'illegible'

    os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)

    out_row = 1
    count = 0
    pred_lines = []
    sample_start = 1
    max_row = ws.max_row

    while sample_start + 6 <= max_row:
        r = sample_start
        dname = ws.cell(row=r, column=1).value

        if dname is None:
            sample_start += ROWS_PER_SAMPLE
            continue

        # Check user PL at row 7 (r+6), col 3
        user_pl = ws.cell(row=r + 6, column=3).value

        if user_pl is not None and str(user_pl).strip() != '':
            # User filled PL → skip
            sample_start += ROWS_PER_SAMPLE
            continue

        # Blank PL → copy all 7 rows to output
        max_col = ws.max_column
        for row_offset in range(7):
            for col in range(1, max_col + 1):
                src_cell = ws.cell(row=r + row_offset, column=col)
                dst_cell = ws_out.cell(row=out_row + row_offset, column=col)
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font = src_cell.font.copy()

        # Extract info for image matching
        img_id = ws.cell(row=r, column=2).value
        pred_cell = ws.cell(row=r + 4, column=3).value  # row 5: pred string
        gt_cell = ws.cell(row=r + 5, column=3).value     # row 6: gt string

        # Record pred for text file
        conf = ws.cell(row=r, column=4).value
        img_id_str = str(int(img_id)) if isinstance(img_id, (float, int)) else str(img_id)
        pred_lines.append((str(pred_cell), str(gt_cell)))

        # Copy matching image
        img_fname = find_image(ERROR_IMG_DIR, str(dname), img_id, pred_cell, gt_cell)
        if img_fname:
            src_path = os.path.join(ERROR_IMG_DIR, img_fname)
            dst_path = os.path.join(OUTPUT_IMG_DIR, img_fname)
            shutil.copy2(src_path, dst_path)

        out_row += ROWS_PER_SAMPLE  # 7 rows + 1 blank
        count += 1
        sample_start += ROWS_PER_SAMPLE

    wb_out.save(OUTPUT_XLSX)

    with open(OUTPUT_TXT, 'w') as f:
        for pred, gt in pred_lines:
            f.write(f'Closest plausible word to "{pred}", when gt is "{gt}", considering there could be label noise.\n')

    with open(OUTPUT_TSV, 'w') as f:
        for pred, gt in pred_lines:
            f.write(f'{pred}\t{gt}\n')

    print(f'Extracted {count} samples without user PL')
    print(f'Excel saved: {OUTPUT_XLSX}')
    print(f'Preds saved: {OUTPUT_TXT}')
    print(f'Images saved: {OUTPUT_IMG_DIR}/')


if __name__ == '__main__':
    main()
